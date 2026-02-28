"""Campaign Analytics Lab — main Streamlit entry point."""
import os
import streamlit as st
import pandas as pd
import numpy as np
from datetime import date

from config import PROFILES_PATH, YEARLY_KPI_PATH, DATASET_PATH, LOGO_PATH
from engine.dna import (
    compute_similarity_weights,
    build_pure_dna,
    build_year_dataframe,
    build_dna_layers,
)
from engine.calibration import calibrate_base, build_projections
from engine.settings_store import load_settings
from views.dashboard import render_dashboard
from views.lab import render_lab
from views.settings import render_settings
from views.docs import render_docs

st.set_page_config(
    page_title="Campaign Analytics Lab",
    layout="wide",
    page_icon="📊",
)

# ─── AUTH CREDENTIALS (env vars with demo fallback) ────────────────────────────
_USER  = os.environ.get("APP_USERNAME", "demo")
_PASS  = os.environ.get("APP_PASSWORD", "demo2026")
_EMAIL = os.environ.get("APP_EMAIL",    "hello@campaign-analytics-lab.io")

_ORANGE = "#F47920"
_BLACK  = "#111111"

# ─── LOGO ──────────────────────────────────────────────────────────────────────
_LOGO_EXISTS = os.path.exists(LOGO_PATH)


def _logo_b64() -> str:
    if not _LOGO_EXISTS:
        return ""
    import base64
    with open(LOGO_PATH, "rb") as f:
        return base64.b64encode(f.read()).decode()

_LOGO_B64 = _logo_b64()

# ─── GLOBAL CSS ────────────────────────────────────────────────────────────────
_CSS = f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

:root {{
    --orange:  {_ORANGE};
    --black:   {_BLACK};
    --bg:      #F8F8F8;
    --surface: #FFFFFF;
    --border:  #EFEFEF;
    --text-1:  #111111;
    --text-2:  #555555;
    --text-3:  #AAAAAA;
    --radius:  10px;
    --shadow:  0 1px 12px rgba(0,0,0,0.06);
    --shadow-hover: 0 4px 20px rgba(0,0,0,0.10);
}}

h1,h2,h3,h4,h5,h6,p,label,
.stMarkdown,.stCaption,
[data-baseweb="tab"],
[data-testid="stMetricLabel"],
[data-testid="stMetricValue"],
[data-testid="stMetricDelta"] {{
    font-family: 'Inter', sans-serif;
}}

[data-testid="stAppViewContainer"] {{ background: var(--bg); }}

@media (min-width: 768px) {{
    [data-testid="stHeader"] {{ display: none !important; }}
}}
@media (max-width: 767px) {{
    [data-testid="stHeader"] {{
        background: var(--surface) !important;
        border-bottom: 1px solid var(--border) !important;
    }}
    [data-testid="stToolbar"] {{ display: none !important; }}
}}

.main .block-container {{
    padding-top: 2.5rem;
    padding-left: 2.5rem;
    padding-right: 2.5rem;
}}
@media (max-width: 767px) {{
    .main .block-container {{
        padding-top: 4rem;
        padding-left: 1rem;
        padding-right: 1rem;
    }}
}}

section[data-testid="stSidebar"],
section[data-testid="stSidebar"] > div:first-child {{
    background: var(--surface);
    border-right: 1px solid var(--border);
}}
section[data-testid="stSidebar"] p           {{ color: var(--text-2); font-size: 0.82rem; }}
section[data-testid="stSidebar"] label       {{ color: var(--text-1); font-size: 0.85rem; }}
section[data-testid="stSidebar"] .stCaption  {{ color: var(--text-3); font-size: 0.78rem; }}
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {{
    color: var(--text-3) !important;
    font-size: 0.6rem !important;
    font-weight: 700 !important;
    text-transform: uppercase;
    letter-spacing: 0.14em;
    margin: 6px 0 2px;
}}
section[data-testid="stSidebar"] hr {{ border-color: var(--border); margin: 8px 0; }}
section[data-testid="stSidebar"] [data-baseweb="radio"] svg {{ fill: var(--orange) !important; }}

[data-baseweb="tab-list"] {{
    background: transparent !important;
    border-radius: 0 !important;
    gap: 0 !important;
    padding: 0 !important;
    border-bottom: 1px solid var(--border) !important;
}}
[data-baseweb="tab"] {{
    border-radius: 0 !important;
    background: transparent !important;
    padding: 10px 20px !important;
    color: var(--text-3) !important;
    font-weight: 500 !important;
    font-size: 0.88rem !important;
    border-bottom: 2px solid transparent !important;
    margin-bottom: -1px !important;
    transition: color 0.15s !important;
}}
[data-baseweb="tab"]:hover {{ color: var(--text-2) !important; }}
[aria-selected="true"] {{
    background: transparent !important;
    color: var(--orange) !important;
    border-bottom: 2px solid var(--orange) !important;
}}

[data-testid="stMetric"] {{
    background: var(--surface);
    border-radius: var(--radius);
    border: 1px solid var(--border);
    border-left: 3px solid var(--orange);
    box-shadow: var(--shadow);
    padding: 20px 22px;
    transition: box-shadow 0.2s, transform 0.2s;
}}
[data-testid="stMetric"]:hover {{
    box-shadow: var(--shadow-hover);
    transform: translateY(-1px);
}}
[data-testid="stMetricLabel"] {{
    color: var(--text-3) !important;
    font-size: 0.65rem !important;
    font-weight: 700 !important;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-bottom: 4px;
}}
[data-testid="stMetricValue"] {{
    color: var(--black) !important;
    font-weight: 700 !important;
    line-height: 1.2 !important;
}}

[data-testid="baseButton-primary"],
[data-testid="baseButton-primaryFormSubmit"] {{
    background: var(--orange) !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    letter-spacing: 0.01em !important;
    transition: opacity 0.15s, transform 0.15s !important;
}}
[data-testid="baseButton-primary"]:hover,
[data-testid="baseButton-primaryFormSubmit"]:hover {{
    opacity: 0.88 !important;
    transform: translateY(-1px) !important;
}}

[data-testid="baseButton-secondary"],
[data-testid="baseButton-secondaryFormSubmit"] {{
    background: #FFFFFF !important;
    color: var(--text-1) !important;
    border: 1px solid #DEDEDE !important;
    border-radius: 8px !important;
    font-weight: 500 !important;
}}
[data-testid="baseButton-secondary"]:hover,
[data-testid="baseButton-secondaryFormSubmit"]:hover {{
    border-color: #BBBBBB !important;
    background: #FAFAFA !important;
}}

[data-testid="stDownloadButton"] > button {{
    background: var(--orange) !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
}}
[data-testid="stDownloadButton"] > button:hover {{ opacity: 0.88 !important; }}

input[type="text"]:focus,
input[type="number"]:focus,
input[type="password"]:focus,
textarea:focus {{
    border-color: var(--orange) !important;
    box-shadow: 0 0 0 2px rgba(244,121,32,0.15) !important;
    outline: none !important;
}}

[data-testid="stFileUploadDropzone"] {{
    border: 2px dashed #E0E0E0;
    border-radius: var(--radius);
    background: #FAFAFA;
}}
[data-testid="stFileUploadDropzone"]:hover {{
    border-color: var(--orange);
    background: #FFF8F4;
}}

[data-testid="stExpander"] {{
    background: var(--surface);
    border: 1px solid var(--border) !important;
    border-radius: var(--radius) !important;
}}

[data-testid="stAlert"] {{ border-radius: 8px; }}
</style>
"""

# ─── AUTH GATE ─────────────────────────────────────────────────────────────────
if "_auth_ok"   not in st.session_state: st.session_state._auth_ok   = False
if "_user_name" not in st.session_state: st.session_state._user_name = ""

if not st.session_state._auth_ok:
    st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
[data-testid="stAppViewContainer"] {{ background: #F8F8F8 !important; }}
[data-testid="stHeader"]  {{ display: none !important; }}
[data-testid="stSidebar"] {{ display: none !important; }}
footer                    {{ display: none !important; }}
.main .block-container {{
    max-width: 420px; padding-top: 8vh;
    padding-left: 1.5rem; padding-right: 1.5rem;
    margin: 0 auto;
}}
[data-testid="stForm"] {{
    background: #FFFFFF; border: 1px solid #EFEFEF;
    border-radius: 16px; padding: 36px 36px 28px;
    box-shadow: 0 4px 32px rgba(0,0,0,0.07);
}}
[data-testid="stTextInput"] > div > div > input {{
    border: 1px solid #E0E0E0; border-radius: 8px;
    color: #111111; padding: 10px 14px;
}}
[data-testid="stTextInput"] > div > div > input:focus {{
    border-color: {_ORANGE}; box-shadow: 0 0 0 2px rgba(244,121,32,0.15); outline: none;
}}
[data-testid="stTextInput"] label {{
    font-family: 'Inter', sans-serif; font-size: 0.8rem;
    font-weight: 500; color: #555555;
}}
[data-testid="baseButton-secondaryFormSubmit"],
[data-testid="baseButton-primaryFormSubmit"] {{
    background: {_ORANGE} !important; color: #FFFFFF !important;
    border: none !important; border-radius: 8px !important;
    font-weight: 600 !important; font-family: 'Inter', sans-serif !important;
    transition: opacity 0.15s !important;
}}
[data-testid="baseButton-secondaryFormSubmit"]:hover,
[data-testid="baseButton-primaryFormSubmit"]:hover {{ opacity: 0.88 !important; }}
</style>
""", unsafe_allow_html=True)

    logo_html = (
        f"<img src='data:image/png;base64,{_LOGO_B64}' "
        f"style='width:72px;border-radius:50%;display:block;margin:0 auto 16px'>"
        if _LOGO_B64 else
        "<div style='font-size:2.5rem;text-align:center;margin-bottom:16px'>📊</div>"
    )
    st.markdown(logo_html, unsafe_allow_html=True)
    st.markdown(
        "<h3 style='text-align:center;font-family:Inter,sans-serif;font-weight:700;"
        "color:#111;margin:0 0 4px;font-size:1.35rem'>Campaign Analytics Lab</h3>"
        "<p style='text-align:center;font-family:Inter,sans-serif;color:#AAAAAA;"
        "font-size:0.84rem;margin-bottom:4px'>Sign in to continue</p>"
        f"<p style='text-align:center;font-family:Inter,sans-serif;color:#CCCCCC;"
        f"font-size:0.75rem;margin-bottom:24px'>"
        f"Demo: <code>demo</code> / <code>demo2026</code></p>",
        unsafe_allow_html=True,
    )

    with st.form("login_form"):
        full_name = st.text_input("Full Name", placeholder="Your name")
        username  = st.text_input("Username")
        password  = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Sign In", use_container_width=True)

    if submitted:
        if not full_name.strip():
            st.error("Please enter your full name.")
        elif username == _USER and password == _PASS:
            st.session_state._auth_ok   = True
            st.session_state._user_name = full_name.strip()
            st.rerun()
        else:
            st.error("Invalid username or password.")

    st.stop()

# ─── INJECT GLOBAL CSS ─────────────────────────────────────────────────────────
st.markdown(_CSS, unsafe_allow_html=True)

# ─── SESSION STATE ─────────────────────────────────────────────────────────────
if "event_log"        not in st.session_state: st.session_state.event_log        = []
if "shock_library"    not in st.session_state: st.session_state.shock_library    = []
if "shift_target_idx" not in st.session_state: st.session_state.shift_target_idx = None
if "tgt_start"        not in st.session_state: st.session_state.tgt_start        = date(2026, 1, 1)
if "tgt_end"          not in st.session_state: st.session_state.tgt_end          = date(2026, 12, 31)
if "target_metric"    not in st.session_state: st.session_state.target_metric    = "Revenue"
if "target_val"       not in st.session_state: st.session_state.target_val       = 100_000.0
if "ui_res_level"     not in st.session_state: st.session_state.ui_res_level     = "Monthly"
if "ui_t_start"       not in st.session_state: st.session_state.ui_t_start       = date(2025, 8, 1)
if "ui_t_end"         not in st.session_state: st.session_state.ui_t_end         = date(2025, 8, 31)
if "ui_s_val"         not in st.session_state: st.session_state.ui_s_val         = 10_000.0
if "ui_conv_val"      not in st.session_state: st.session_state.ui_conv_val      = 200.0
if "ui_rev_val"       not in st.session_state: st.session_state.ui_rev_val       = 20_000.0
if "ui_adj_s"         not in st.session_state: st.session_state.ui_adj_s         = 0.0
if "ui_adj_conv"      not in st.session_state: st.session_state.ui_adj_conv      = 0.0
if "ui_adj_rev"       not in st.session_state: st.session_state.ui_adj_rev       = 0.0
if "ui_sel_brands"    not in st.session_state: st.session_state.ui_sel_brands    = []

# ─── DATA LOADING ──────────────────────────────────────────────────────────────
@st.cache_data
def load_data():
    profiles    = pd.read_csv(PROFILES_PATH)
    profiles["Year"] = profiles["Year"].astype(str)
    yearly_kpis = pd.read_csv(YEARLY_KPI_PATH)
    df_raw      = pd.read_csv(DATASET_PATH)
    df_raw["Date"]  = pd.to_datetime(df_raw["Date"])
    df_raw["brand"] = df_raw["brand"].str.strip().str.lower()
    return profiles, yearly_kpis, df_raw

profiles, yearly_kpis, df_raw = load_data()

data_years  = sorted([int(y) for y in profiles["Year"].unique() if y != "Overall"])
min_data_yr = data_years[0]
max_data_yr = data_years[-1]

_settings = load_settings()

# ─── SIDEBAR ───────────────────────────────────────────────────────────────────
logo_html = (
    f"<img src='data:image/png;base64,{_LOGO_B64}' "
    f"style='width:72px;border-radius:50%;display:inline-block'>"
    if _LOGO_B64 else
    "<div style='width:72px;height:72px;border-radius:50%;background:#F8F8F8;"
    "display:inline-flex;align-items:center;justify-content:center;"
    "font-size:2rem'>📊</div>"
)
st.sidebar.markdown(
    f"<div style='text-align:center;padding:28px 16px 16px'>"
    f"{logo_html}"
    f"<div style='margin-top:10px;font-family:Inter,sans-serif;font-weight:700;"
    f"font-size:0.9rem;color:#111111;letter-spacing:-0.01em'>Campaign Analytics Lab</div>"
    f"<div style='font-family:Inter,sans-serif;font-size:0.68rem;color:#AAAAAA;"
    f"margin-top:2px'>Decision Intelligence Platform</div>"
    f"<div style='font-family:Inter,sans-serif;font-size:0.72rem;color:#555555;"
    f"margin-top:6px;font-weight:500'>{st.session_state._user_name}</div>"
    f"</div>",
    unsafe_allow_html=True,
)
st.sidebar.divider()

# ── Navigation ─────────────────────────────────────────────────────────────────
_NAV_LABELS = [
    "📊 Dashboard",
    "⚡ Simulation Lab",
    "⚙️ Settings",
    "📖 Documentation",
]

if "nav_page_idx" not in st.session_state:
    st.session_state.nav_page_idx = 0

_page_idx = st.sidebar.radio(
    "Navigate",
    options=list(range(len(_NAV_LABELS))),
    format_func=lambda i: _NAV_LABELS[i],
    key="nav_page_idx",
    label_visibility="collapsed",
)
page = _NAV_LABELS[_page_idx]
st.sidebar.divider()

_is_analytics = page in ("📊 Dashboard", "⚡ Simulation Lab")

# ── Analytics-only sidebar controls ────────────────────────────────────────────
if _is_analytics:
    st.sidebar.header("Market Resolution")
    res_level = st.sidebar.radio(
        "Granularity", ["Monthly", "Weekly", "Daily"], key="ui_res_level")
    time_col = ("Month" if res_level == "Monthly"
                else "Week" if res_level == "Weekly" else "DayOfYear")

    st.sidebar.markdown(
        "<p style='font-size:0.75rem;font-weight:600;color:#555;margin:8px 0 2px;"
        "font-family:Inter,sans-serif'>Entities</p>",
        unsafe_allow_html=True,
    )
    all_brands = sorted(profiles["brand"].unique())
    if not st.session_state.ui_sel_brands:
        st.session_state.ui_sel_brands = list(all_brands)
    select_all = st.sidebar.checkbox("All entities", value=True)
    sel_brands = []
    for b in all_brands:
        chk = st.sidebar.checkbox(
            b.title(), value=(select_all or b in st.session_state.ui_sel_brands),
            disabled=select_all, key=f"chk_{b}")
        if select_all or chk:
            sel_brands.append(b)
    st.session_state.ui_sel_brands = sel_brands

    if not sel_brands:
        st.warning("Select at least one entity.")
        st.stop()

    st.sidebar.divider()
    st.sidebar.header("Trial Reality")
    t_start  = st.sidebar.date_input("Start Date", key="ui_t_start")
    t_end    = st.sidebar.date_input("End Date",   key="ui_t_end")
    s_val    = st.sidebar.number_input("Sessions",    min_value=0.0, key="ui_s_val")
    conv_val = st.sidebar.number_input("Conversions", min_value=0.0, key="ui_conv_val")
    rev_val  = st.sidebar.number_input("Revenue (€)", min_value=0.0, key="ui_rev_val")

    if t_start.year < min_data_yr or t_end.year > max_data_yr + 2:
        st.sidebar.warning(f"Outside data range ({min_data_yr}–{max_data_yr}).")

    with st.sidebar.expander("Pre-Adjustment"):
        st.caption(
            "+ % = trial was boosted (strip lift back). "
            "− % = trial was suppressed (add lift back).")
        st.number_input("Sessions adj (%)",    -100.0, 500.0, key="ui_adj_s",    step=5.0)
        st.number_input("Conversions adj (%)", -100.0, 500.0, key="ui_adj_conv", step=5.0)
        st.number_input("Revenue adj (%)",     -100.0, 500.0, key="ui_adj_rev",  step=5.0)

    adj_s    = s_val    / (1 + st.session_state.ui_adj_s    / 100) if (1 + st.session_state.ui_adj_s    / 100) != 0 else s_val
    adj_conv = conv_val / (1 + st.session_state.ui_adj_conv / 100) if (1 + st.session_state.ui_adj_conv / 100) != 0 else conv_val
    adj_rev  = rev_val  / (1 + st.session_state.ui_adj_rev  / 100) if (1 + st.session_state.ui_adj_rev  / 100) != 0 else rev_val

    proj_year    = str(t_start.year)
    norm_weights = compute_similarity_weights(
        profiles, sel_brands, proj_year, t_start, t_end, s_val, conv_val, rev_val)

    st.sidebar.divider()
    st.sidebar.header("DNA Weights")
    st.sidebar.caption("35% — All-time overall")
    for y, w in norm_weights.items():
        st.sidebar.caption(f"{w * 65.0:.1f}% — {y}")

    pure_dna       = build_pure_dna(profiles, sel_brands, norm_weights)
    df, _full_year = build_year_dataframe(int(proj_year))
    build_dna_layers(df, pure_dna, st.session_state.event_log)

    base_sessions, base_cr, base_aov = calibrate_base(
        df, t_start, t_end, adj_s, adj_conv, adj_rev)
    if base_sessions is None:
        st.error("Trial date range yields zero DNA sum. Widen the trial period.")
        st.stop()

    build_projections(df, base_sessions, base_cr, base_aov, st.session_state.event_log)

    st.sidebar.divider()
    st.sidebar.header("Export")

    # Build a simple Excel export
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import openpyxl.utils

        def _build_excel(df_exp):
            wb = Workbook()
            ws = wb.active
            ws.title = "Projections"
            cols = [
                "Date", "Sessions_Base", "Sessions_Sim",
                "Conversions_Base", "Conversions_Sim",
                "Revenue_Base", "Revenue_Sim",
            ]
            hdr = Font(bold=True, color="FFFFFF")
            fill = PatternFill("solid", fgColor="1A1A6B")
            for ci, c in enumerate(cols, 1):
                cell = ws.cell(1, ci, c)
                cell.font = hdr
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")
            for ri, row in df_exp[cols].iterrows():
                for ci, v in enumerate(row, 1):
                    ws.cell(ri + 2, ci, v)
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        xl = _build_excel(df)
        st.sidebar.download_button(
            "Download Strategy Report",
            data=xl,
            file_name=f"strategy_{proj_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception:
        pass

# ── Sidebar footer ──────────────────────────────────────────────────────────────
st.sidebar.divider()
st.sidebar.markdown(
    f"<p style='font-size:0.72rem;color:#CCCCCC;font-family:Inter,sans-serif;"
    f"text-align:center;margin:0 0 8px'>"
    f"<a href='mailto:{_EMAIL}' style='color:{_ORANGE};text-decoration:none'>"
    f"{_EMAIL}</a></p>",
    unsafe_allow_html=True,
)
if st.sidebar.button("Sign Out", use_container_width=True):
    st.session_state._auth_ok   = False
    st.session_state._user_name = ""
    st.rerun()

# ─── PAGE HEADER ───────────────────────────────────────────────────────────────
_subtitles = {
    "📊 Dashboard":      "Forecast projection, goal tracking, and demand DNA profile",
    "⚡ Simulation Lab": "Campaign simulation, DNA sculpting, de-shock analysis, attribution",
    "⚙️ Settings":       "Campaign default coefficients per entity",
    "📖 Documentation":  "Models, formulas, assumptions, and complete usage guide",
}
st.markdown(
    f"<div style='margin-bottom:24px'>"
    f"<h1 style='font-family:Inter,sans-serif;font-weight:700;font-size:1.6rem;"
    f"color:{_BLACK};margin:0 0 4px;letter-spacing:-0.025em'>{page}</h1>"
    f"<p style='font-family:Inter,sans-serif;font-size:0.84rem;color:#AAAAAA;"
    f"margin:0;font-weight:400'>{_subtitles.get(page,'')}</p>"
    f"<div style='height:1px;background:#EBEBEB;margin-top:16px'></div></div>",
    unsafe_allow_html=True,
)

# ─── PAGE ROUTING ──────────────────────────────────────────────────────────────
if page == "📊 Dashboard":
    render_dashboard(
        df, profiles, yearly_kpis,
        sel_brands, res_level, time_col,
        base_cr, base_aov,
    )
elif page == "⚡ Simulation Lab":
    render_lab(
        df, df_raw, sel_brands, res_level, time_col,
        base_sessions, base_cr, base_aov,
        adj_s, adj_conv, adj_rev,
        t_start, t_end, pure_dna,
        settings=_settings,
    )
elif page == "⚙️ Settings":
    render_settings()
elif page == "📖 Documentation":
    render_docs()
